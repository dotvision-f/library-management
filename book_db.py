from openpyxl import Workbook
from openpyxl import load_workbook

class Book_DB:

    def __init__(self):
        self.header = ['title', 'author', 'subject', 'PDF link']
        self.book_db = Workbook()
        self.sheet = self.book_db.active
        self.count = 1

    def write_header(self):
        self.sheet["A1"] = self.header[0]
        self.sheet["B1"] = self.header[1]
        self.sheet["C1"] = self.header[2]
        self.sheet["D1"] = self.header[3]

    def save_db(self):
        self.book_db.save("Book_Database.xlsx")

    def add_book(self, title, author, subject, PDF_link):
        self.sheet[f"A{self.count + 1}"] = title
        self.sheet[f"B{self.count + 1}"] = author
        self.sheet[f"C{self.count + 1}"] = subject
        self.sheet[f"D{self.count + 1}"] = PDF_link
        self.count += 1
        self.save_db()

    def remove_book(self, title):
        row_index = self.find_row_index(title)
        if row_index is not None:
            for cell in self.sheet[row_index]:
                cell.value = None
            self.count -= 1
        else:
            print("Cannot find the book. Please enter the right book title!")
        self.save_db()

    def find_row_index(self, word):
        word = word
        for row_index, row in enumerate(self.sheet.iter_rows()):
            for cell in row:
                if cell.value == word:
                    return row_index + 1
        return None

    def search_book_title(self, title):     #tra ve list sach
        book_list = []
        for i in range(1, self.count+1):
            if str(title) in str(self.sheet[f"A{i}"].value):
                book_list.append(self.sheet[f"A{i}"].value)
        if len(book_list) >= 1:
            return book_list
        else:
            return []

    def search_book_author(self, author):
        book_list = []
        for i in range(1, self.count+1):
            if str(self.sheet[f"B{i}"].value) == str(author):
                book_list.append(self.sheet[f"A{i}"].value)
        if len(book_list) >= 1:
            return book_list
        else:
            return []

    def search_book_subject(self, subject):
        book_list = []
        for i in range(1, self.count+1):
            if str(subject) in str(self.sheet[f"C{i}"].value) :
                book_list.append(self.sheet[f"A{i}"].value)
        if len(book_list) >= 1:
            return book_list
        else:
            return []
    
    def change_info(self, title, type_info_to_change, new_info):
        row_index = self.find_row_index(title)
        if str(type_info_to_change) == 'title':
            self.sheet[f"A{row_index}"].value = new_info
        elif str(type_info_to_change) == 'author':
            self.sheet[f"B{row_index}"].value = new_info
        elif str(type_info_to_change) == 'subject':
            self.sheet[f"C{row_index}"].value = new_info
        elif str(type_info_to_change) == 'PDF_link':
            self.sheet[f"D{row_index}"].value = new_info
        else:
            print("please enter the correct type of data to change: email, password or in bag.")


# book_db = Book_DB()
# book_db.write_header()
# book_db.add_book('anh yeu em', 'dung le', 'poetry', './desktop/anhyeuem.pdf')
# print(book_db.search_book_author('dung le'))

# book_db.save_db()