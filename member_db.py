from openpyxl import Workbook
from openpyxl import load_workbook

class Member_DB:

    def __init__(self):
        self.header = ['email', 'password', 'level', 'in bag']
        self.mem_db = Workbook()
        self.sheet = self.mem_db.active
        self.count = 1

    def write_header(self):         #dau tien, goi ham nay de khoi tao database
        self.sheet["A1"] = self.header[0]
        self.sheet["B1"] = self.header[1]
        self.sheet["C1"] = self.header[2]
        self.sheet["D1"] = self.header[3]

    def save_db(self):              #sau khi thao tac xong, goi ham nay de luu database
        self.mem_db.save("Member_Database.xlsx")

    def add_member(self, email, password):
        email = str(email)
        password = str(password)
        self.sheet[f"A{self.count + 1}"] = email
        self.sheet[f"B{self.count + 1}"] = password
        self.sheet[f"D{self.count + 1}"] = None

        email_split = email.split("@")
        if email_split[1] == 'st.phenikaa-uni.edu.vn':
            self.sheet[f"C{self.count + 1}"] = "Student"
        elif email_split[1] == 'phenikaa-uni.edu.vn':
            self.sheet[f"C{self.count + 1}"] = "Teacher"
        else:
            self.sheet[f"C{self.count + 1}"] = "Guest"

        self.count += 1

    def remove_member(self, email):
        email = email
        row_index = self.find_row_index(email)
        if row_index is not None:
            for cell in self.sheet[row_index]:
                cell.value = None
            self.count -= 1
        else:
            print("Cannot find the email. Please enter the right email!")

    def find_row_index(self, word):
        word = word
        for row_index, row in enumerate(self.sheet.iter_rows()):
            for cell in row:
                if cell.value == word:
                    return row_index + 1
        return None

    def add_to_bag(self, email, book_name):         #them sach vao gio hang
        email = email
        book_name = book_name
        row_index = self.find_row_index(email)
        column_d_value = self.sheet[f'D{row_index}']
        column_d_value.value = book_name

    def show_info(self, email):
        row_index = self.find_row_index(email)
        print("email: ", self.sheet[f"A{row_index}"].value)
        print("password: ", self.sheet[f"B{row_index}"].value)
        print("level: ", self.sheet[f"C{row_index}"].value)
        print("in bag: ", self.sheet[f"D{row_index}"].value)

    def access_db(self):
        workbook = load_workbook(filename = "Member_Database.xlsx")
        
    def change_info(self, email, type_info_to_change, new_info):
        row_index = self.find_row_index(email)
        if str(type_info_to_change) == 'email':
            self.sheet[f"A{row_index}"].value = new_info
        elif str(type_info_to_change) == 'password':
            self.sheet[f"B{row_index}"].value = new_info
        elif str(type_info_to_change) == 'in bag':
            self.sheet[f"D{row_index}"].value = new_info
        else:
            print("please enter the correct type of data to change: email, password or in bag.")

#Mem_db = Member_DB()
#Mem_db.write_header()
#Mem_db.add_member('quangdung@gmail.com', 21032002)
#Mem_db.add_member('dungle@gmail.com', 32323232)
#Mem_db.remove_member('dungle@gmail.com')
#Mem_db.show_info('quangdung@gmail.com')

#xong roi thi lu file excel bang cai nay
#Mem_db.save_db()